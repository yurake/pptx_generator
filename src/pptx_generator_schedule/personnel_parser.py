"""開発要員計画のxlsxファイルをパースしてJSONを生成する。

インプットファイル2（xlsx）から中間生成ファイル1（JSON）を作成し、
さらにインプットファイル1（schedule.md）と統合して中間生成ファイル2を作成する。
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from .models import (
    DevelopmentPersonnelPlan,
    DisplayUnit,
    PersonnelData,
    PersonnelMessage,
    PersonnelMonthData,
    PersonnelMonthSummary,
    PersonnelPhaseMonthSummary,
    PersonnelPhaseSummary,
    PersonnelPhaseData,
    PersonnelQuarterSummary,
    PersonnelScheduleMilestone,
    PersonnelSchedulePhase,
    PersonnelScheduleTask,
    PersonnelTaskData,
    ScheduleGantt,
)
from .schedule_parser import parse_schedule_markdown


def parse_personnel_xlsx(xlsx_path: str | Path) -> PersonnelData:
    """xlsxファイルから工数データをパースする。

    Args:
        xlsx_path: 工数xlsxファイルのパス

    Returns:
        PersonnelData: パースされた工数データ（中間生成ファイル1）
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        msg = f"xlsxファイルが見つかりません: {xlsx_path}"
        raise FileNotFoundError(msg)

    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    # ヘッダー行を解析して月の列を特定
    month_columns = _parse_header(ws)

    # データ行をパース
    phases: dict[str, PersonnelPhaseData] = {}

    for row_idx in range(4, ws.max_row + 1):
        phase_name = ws.cell(row=row_idx, column=1).value
        task_name = ws.cell(row=row_idx, column=2).value

        if not phase_name or not task_name:
            continue

        # フェーズを取得または作成
        if phase_name not in phases:
            phases[phase_name] = PersonnelPhaseData(phase_name=phase_name, tasks=[])

        # 月別工数データを収集
        months_data: list[PersonnelMonthData] = []
        for (year, month), (emp_col, pn_col, si_col) in month_columns.items():
            emp_val = ws.cell(row=row_idx, column=emp_col).value
            pn_val = ws.cell(row=row_idx, column=pn_col).value
            si_val = ws.cell(row=row_idx, column=si_col).value

            # 値が存在する場合のみ追加
            if any(v is not None and v != 0 for v in [emp_val, pn_val, si_val]):
                months_data.append(
                    PersonnelMonthData(
                        year=year,
                        month=month,
                        employee=float(emp_val or 0),
                        pn=float(pn_val or 0),
                        si=float(si_val or 0),
                    )
                )

        # タスクを追加
        task = PersonnelTaskData(task_name=task_name, months=months_data)
        phases[phase_name].tasks.append(task)

    return PersonnelData(
        extracted_at=datetime.now(timezone.utc).isoformat(),
        source_path=str(xlsx_path),
        phases=list(phases.values()),
    )


def _parse_header(ws) -> dict[tuple[int, int], tuple[int, int, int]]:
    """ヘッダー行を解析して月→列番号のマッピングを返す。

    Returns:
        dict: {(year, month): (emp_col, pn_col, si_col)}
    """
    month_columns: dict[tuple[int, int], tuple[int, int, int]] = {}

    # 行1: 年度ヘッダー、行2: 月ヘッダー、行3: 工数種別
    col = 3
    while col <= ws.max_column:
        # 年度を取得（マージセルの場合は最初のセルから取得）
        year_cell = ws.cell(row=1, column=col)
        year_value = year_cell.value
        if year_value is None:
            # マージセルの場合、前の列から年度を継承
            for prev_col in range(col - 1, 2, -1):
                prev_value = ws.cell(row=1, column=prev_col).value
                if prev_value is not None:
                    year_value = prev_value
                    break

        if year_value is None:
            col += 3
            continue

        # 年度を数値に変換
        if isinstance(year_value, str) and "年度" in year_value:
            year = int(year_value.replace("年度", ""))
        elif isinstance(year_value, (int, float)):
            year = int(year_value)
        else:
            col += 3
            continue

        # 月を取得
        month_cell = ws.cell(row=2, column=col)
        month_value = month_cell.value
        if month_value is None:
            col += 3
            continue

        if isinstance(month_value, str) and "月" in month_value:
            month = int(month_value.replace("月", ""))
        elif isinstance(month_value, (int, float)):
            month = int(month_value)
        else:
            col += 3
            continue

        # 実際の年を計算（年度から）
        # 4-12月は年度と同じ年、1-3月は年度+1年
        if 4 <= month <= 12:
            actual_year = year
        else:
            actual_year = year + 1

        month_columns[(actual_year, month)] = (col, col + 1, col + 2)
        col += 3

    return month_columns


def save_personnel_data(data: PersonnelData, output_path: str | Path) -> None:
    """PersonnelDataをJSONファイルとして保存する。"""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        data.model_dump_json(indent=2),
        encoding="utf-8",
    )


def integrate_personnel_schedule(
    schedule_data: ScheduleGantt,
    personnel_data: PersonnelData,
    *,
    schedule_source: str = "",
    personnel_source: str = "",
) -> DevelopmentPersonnelPlan:
    """スケジュールデータと工数データを統合する。

    Args:
        schedule_data: スケジュールデータ（schedule.mdからパース済み）
        personnel_data: 工数データ（xlsxからパース済み）
        schedule_source: スケジュールファイルのパス
        personnel_source: 工数ファイルのパス

    Returns:
        DevelopmentPersonnelPlan: 統合されたデータ（中間生成ファイル2）
    """
    # タスク名でスケジュールと工数をマッチング
    schedule_tasks: dict[str, tuple[str, str]] = {}  # task_name -> (start, end)
    for project in schedule_data.projects:
        for task in project.tasks:
            schedule_tasks[task.name] = (task.start, task.end)

    # フェーズ名でスケジュールと工数をマッチング
    phase_mapping: dict[str, str] = {}  # personnel phase -> schedule project
    for phase in personnel_data.phases:
        for project in schedule_data.projects:
            # フェーズ名の一部がプロジェクト名に含まれているか確認
            if project.name in phase.phase_name or phase.phase_name in project.name:
                phase_mapping[phase.phase_name] = project.name
                break

    # 統合フェーズデータを構築
    phases: list[PersonnelSchedulePhase] = []
    for phase in personnel_data.phases:
        tasks: list[PersonnelScheduleTask] = []
        for task in phase.tasks:
            # スケジュールからタスクの開始・終了日を取得
            start_date, end_date = schedule_tasks.get(task.task_name, ("", ""))
            if not start_date or not end_date:
                # スケジュールにタスクがない場合、工数データから推定
                if task.months:
                    sorted_months = sorted(task.months, key=lambda m: (m.year, m.month))
                    first = sorted_months[0]
                    last = sorted_months[-1]
                    start_date = f"{first.year}-{first.month:02d}-01"
                    end_date = f"{last.year}-{last.month:02d}-01"
                else:
                    continue

            tasks.append(
                PersonnelScheduleTask(
                    task_name=task.task_name,
                    start_date=start_date,
                    end_date=end_date,
                    months=task.months,
                )
            )

        if tasks:
            phases.append(
                PersonnelSchedulePhase(
                    phase_name=phase.phase_name,
                    tasks=tasks,
                )
            )

    # マイルストーンを変換
    milestones: list[PersonnelScheduleMilestone] = [
        PersonnelScheduleMilestone(name=m.name, date=m.date)
        for m in schedule_data.meta.milestones
    ]

    # 年度一覧を計算
    fiscal_years = _calculate_fiscal_years(personnel_data)

    # 対象月リストと総月数を計算
    target_months = _calculate_target_months(personnel_data)
    total_months = len(target_months)

    # 表示単位を決定（12ヶ月以内なら月単位、超えたら四半期単位）
    display_unit: DisplayUnit = "month" if total_months <= 12 else "quarter"

    # フェーズ別・四半期別サマリを計算（四半期単位用）
    phase_summaries = _calculate_phase_summaries(personnel_data, fiscal_years)

    # 全体四半期サマリを計算
    total_summary = _calculate_total_summary(phase_summaries, fiscal_years)

    # フェーズ別・月別サマリを計算（月単位用）
    phase_month_summaries = _calculate_phase_month_summaries(
        personnel_data, target_months
    )

    # 全体月サマリを計算
    total_month_summary = _calculate_total_month_summary(
        phase_month_summaries, target_months
    )

    # メッセージを生成
    messages = _generate_messages(phase_summaries, schedule_data, milestones)

    return DevelopmentPersonnelPlan(
        generated_at=datetime.now(timezone.utc).isoformat(),
        schedule_source=schedule_source,
        personnel_source=personnel_source,
        title="開発要員計画",
        department="情報システム部",
        display_unit=display_unit,
        total_months=total_months,
        fiscal_years=fiscal_years,
        target_months=target_months,
        phases=phases,
        milestones=milestones,
        phase_summaries=phase_summaries,
        total_summary=total_summary,
        phase_month_summaries=phase_month_summaries,
        total_month_summary=total_month_summary,
        messages=messages,
    )


def _calculate_fiscal_years(personnel_data: PersonnelData) -> list[int]:
    """工数データから対象年度一覧を計算する。"""
    fiscal_years: set[int] = set()
    for phase in personnel_data.phases:
        for task in phase.tasks:
            for month in task.months:
                fiscal_years.add(month.fiscal_year)
    return sorted(fiscal_years)


def _calculate_target_months(personnel_data: PersonnelData) -> list[tuple[int, int]]:
    """工数データから対象月リストを計算する。

    Returns:
        list[tuple[int, int]]: (年, 月)のリスト（年度順でソート）
    """
    months_set: set[tuple[int, int]] = set()
    for phase in personnel_data.phases:
        for task in phase.tasks:
            for month in task.months:
                months_set.add((month.year, month.month))

    # 年度順でソート（4月始まり）
    def sort_key(ym: tuple[int, int]) -> tuple[int, int]:
        year, month = ym
        # 年度を計算
        fiscal_year = year if month >= 4 else year - 1
        # 年度内の月順（4月=0, 5月=1, ... 3月=11）
        month_in_fy = (month - 4) % 12
        return (fiscal_year, month_in_fy)

    return sorted(months_set, key=sort_key)


def _calculate_phase_summaries(
    personnel_data: PersonnelData,
    fiscal_years: list[int],
) -> list[PersonnelPhaseSummary]:
    """フェーズ別・四半期別サマリを計算する。"""
    summaries: list[PersonnelPhaseSummary] = []

    for phase in personnel_data.phases:
        # 四半期別に集計
        quarter_data: dict[tuple[int, int], dict[str, float]] = {}
        for fy in fiscal_years:
            for q in range(1, 5):
                quarter_data[(fy, q)] = {"employee": 0.0, "pn": 0.0, "si": 0.0}

        total_emp = 0.0
        total_pn = 0.0
        total_si = 0.0

        for task in phase.tasks:
            for month in task.months:
                fy = month.fiscal_year
                q = month.quarter
                if (fy, q) in quarter_data:
                    quarter_data[(fy, q)]["employee"] += month.employee
                    quarter_data[(fy, q)]["pn"] += month.pn
                    quarter_data[(fy, q)]["si"] += month.si
                total_emp += month.employee
                total_pn += month.pn
                total_si += month.si

        quarters = [
            PersonnelQuarterSummary(
                fiscal_year=fy,
                quarter=q,
                employee=quarter_data[(fy, q)]["employee"],
                pn=quarter_data[(fy, q)]["pn"],
                si=quarter_data[(fy, q)]["si"],
            )
            for fy, q in sorted(quarter_data.keys())
        ]

        summaries.append(
            PersonnelPhaseSummary(
                phase_name=phase.phase_name,
                quarters=quarters,
                total_employee=total_emp,
                total_pn=total_pn,
                total_si=total_si,
            )
        )

    return summaries


def _calculate_total_summary(
    phase_summaries: list[PersonnelPhaseSummary],
    fiscal_years: list[int],
) -> list[PersonnelQuarterSummary]:
    """全体の四半期別サマリを計算する。"""
    quarter_totals: dict[tuple[int, int], dict[str, float]] = {}
    for fy in fiscal_years:
        for q in range(1, 5):
            quarter_totals[(fy, q)] = {"employee": 0.0, "pn": 0.0, "si": 0.0}

    for phase in phase_summaries:
        for quarter in phase.quarters:
            key = (quarter.fiscal_year, quarter.quarter)
            if key in quarter_totals:
                quarter_totals[key]["employee"] += quarter.employee
                quarter_totals[key]["pn"] += quarter.pn
                quarter_totals[key]["si"] += quarter.si

    return [
        PersonnelQuarterSummary(
            fiscal_year=fy,
            quarter=q,
            employee=quarter_totals[(fy, q)]["employee"],
            pn=quarter_totals[(fy, q)]["pn"],
            si=quarter_totals[(fy, q)]["si"],
        )
        for fy, q in sorted(quarter_totals.keys())
    ]


def _calculate_phase_month_summaries(
    personnel_data: PersonnelData,
    target_months: list[tuple[int, int]],
) -> list[PersonnelPhaseMonthSummary]:
    """フェーズ別・月別サマリを計算する（月単位表示用）。"""
    summaries: list[PersonnelPhaseMonthSummary] = []

    for phase in personnel_data.phases:
        # 月別に集計
        month_data: dict[tuple[int, int], dict[str, float]] = {}
        for ym in target_months:
            month_data[ym] = {"employee": 0.0, "pn": 0.0, "si": 0.0}

        total_emp = 0.0
        total_pn = 0.0
        total_si = 0.0

        for task in phase.tasks:
            for month in task.months:
                key = (month.year, month.month)
                if key in month_data:
                    month_data[key]["employee"] += month.employee
                    month_data[key]["pn"] += month.pn
                    month_data[key]["si"] += month.si
                total_emp += month.employee
                total_pn += month.pn
                total_si += month.si

        months = [
            PersonnelMonthSummary(
                year=year,
                month=month,
                employee=month_data[(year, month)]["employee"],
                pn=month_data[(year, month)]["pn"],
                si=month_data[(year, month)]["si"],
            )
            for year, month in target_months
        ]

        summaries.append(
            PersonnelPhaseMonthSummary(
                phase_name=phase.phase_name,
                months=months,
                total_employee=total_emp,
                total_pn=total_pn,
                total_si=total_si,
            )
        )

    return summaries


def _calculate_total_month_summary(
    phase_month_summaries: list[PersonnelPhaseMonthSummary],
    target_months: list[tuple[int, int]],
) -> list[PersonnelMonthSummary]:
    """全体の月別サマリを計算する（月単位表示用）。"""
    month_totals: dict[tuple[int, int], dict[str, float]] = {}
    for ym in target_months:
        month_totals[ym] = {"employee": 0.0, "pn": 0.0, "si": 0.0}

    for phase in phase_month_summaries:
        for month in phase.months:
            key = (month.year, month.month)
            if key in month_totals:
                month_totals[key]["employee"] += month.employee
                month_totals[key]["pn"] += month.pn
                month_totals[key]["si"] += month.si

    return [
        PersonnelMonthSummary(
            year=year,
            month=month,
            employee=month_totals[(year, month)]["employee"],
            pn=month_totals[(year, month)]["pn"],
            si=month_totals[(year, month)]["si"],
        )
        for year, month in target_months
    ]


def _format_date_as_fiscal_year_month(date_str: str) -> str:
    """日付を年度月形式に変換する。

    Args:
        date_str: YYYY-MM-DD形式の日付文字列

    Returns:
        「YYYY年度M月」形式の文字列
    """
    date = datetime.strptime(date_str, "%Y-%m-%d")
    # 年度を計算（4月始まり）
    fiscal_year = date.year if date.month >= 4 else date.year - 1
    return f"{fiscal_year}年度{date.month}月"


def _generate_messages(
    phase_summaries: list[PersonnelPhaseSummary],
    schedule_data: ScheduleGantt,
    milestones: list[PersonnelScheduleMilestone],
) -> list[PersonnelMessage]:
    """要員計画のポイントメッセージを生成する。"""
    messages: list[PersonnelMessage] = []

    # メッセージ1: 全体の要員計画概要
    total_emp = sum(p.total_employee for p in phase_summaries)
    total_pn = sum(p.total_pn for p in phase_summaries)
    total_si = sum(p.total_si for p in phase_summaries)

    messages.append(
        PersonnelMessage(
            number=1,
            text=f"開発期間全体で社員{total_emp:.1f}人月、PN{total_pn:.1f}人月、SI{total_si:.1f}人月の体制で推進。",
            highlight_quarters=[],
        )
    )

    # メッセージ2: 社員の注力ポイント
    # 設計フェーズを特定
    design_phases = [
        p
        for p in phase_summaries
        if "設計" in p.phase_name or "企画" in p.phase_name
    ]
    if design_phases:
        design_emp = sum(p.total_employee for p in design_phases)
        messages.append(
            PersonnelMessage(
                number=2,
                text=f"社員は設計工程（計{design_emp:.1f}人月）に注力し、品質確保を図る。",
                highlight_quarters=[],
            )
        )
    else:
        messages.append(
            PersonnelMessage(
                number=2,
                text="社員は要件定義・設計工程に注力し、品質確保を図る。",
                highlight_quarters=[],
            )
        )

    # メッセージ3: マイルストーン情報（年度月形式で表記）
    if milestones:
        milestone_texts = [
            f"{m.name}（{_format_date_as_fiscal_year_month(m.date)}）"
            for m in milestones[:3]
        ]
        messages.append(
            PersonnelMessage(
                number=3,
                text=f"主要マイルストーン: {', '.join(milestone_texts)}",
                highlight_quarters=[],
            )
        )

    return messages


def save_development_personnel_plan(
    plan: DevelopmentPersonnelPlan,
    output_path: str | Path,
) -> None:
    """DevelopmentPersonnelPlanをJSONファイルとして保存する。"""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        plan.model_dump_json(indent=2),
        encoding="utf-8",
    )


def create_personnel_plan_from_files(
    schedule_md_path: str | Path,
    personnel_xlsx_path: str | Path,
    output_dir: str | Path,
) -> tuple[PersonnelData, DevelopmentPersonnelPlan]:
    """ファイルからPersonnelDataとDevelopmentPersonnelPlanを生成する。

    Args:
        schedule_md_path: schedule.mdファイルのパス
        personnel_xlsx_path: 工数xlsxファイルのパス
        output_dir: 出力ディレクトリ

    Returns:
        (PersonnelData, DevelopmentPersonnelPlan): 中間生成ファイル1と2のデータ
    """
    schedule_md_path = Path(schedule_md_path)
    personnel_xlsx_path = Path(personnel_xlsx_path)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # schedule.mdをパース
    schedule_data = parse_schedule_markdown(schedule_md_path)

    # xlsxをパースして中間生成ファイル1を作成
    personnel_data = parse_personnel_xlsx(personnel_xlsx_path)
    personnel_json_path = output_dir / "personnel_data.json"
    save_personnel_data(personnel_data, personnel_json_path)

    # スケジュールと工数を統合して中間生成ファイル2を作成
    plan = integrate_personnel_schedule(
        schedule_data,
        personnel_data,
        schedule_source=str(schedule_md_path),
        personnel_source=str(personnel_xlsx_path),
    )
    plan_json_path = output_dir / "development_personnel_plan.json"
    save_development_personnel_plan(plan, plan_json_path)

    return personnel_data, plan