"""AI対応の開発要員計画xlsxパーサー。

Anthropic Claude APIまたはAmazon Bedrock経由でClaudeを使用して、
様々なフォーマットのxlsxファイルから標準化されたPersonnelData JSONを生成する。
"""

from __future__ import annotations

import json
import logging
import os
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Literal

from openpyxl import load_workbook

from personnel.models import (
    PersonnelData,
    PersonnelMonthData,
    PersonnelPhaseData,
    PersonnelTaskData,
)

logger = logging.getLogger(__name__)

_AI_LOGGER = logging.getLogger("executive_board.personnel_ai_parser")

# Anthropic直接API用のデフォルトモデル
DEFAULT_MODEL_ANTHROPIC = "claude-sonnet-4-20250514"
# Amazon Bedrock用のデフォルトモデル (Claude Sonnet 4 - US cross-region inference profile)
DEFAULT_MODEL_BEDROCK = "us.anthropic.claude-sonnet-4-20250514-v1:0"
DEFAULT_MAX_TOKENS = 8192


class PersonnelAIParserError(RuntimeError):
    """AI パーサーのエラー。"""


class PersonnelAIParserConfigurationError(PersonnelAIParserError):
    """AI パーサーの設定エラー。"""


def _xlsx_to_text(xlsx_path: Path) -> str:
    """xlsxファイルの内容をテキスト形式に変換する。
    
    Args:
        xlsx_path: xlsxファイルのパス
        
    Returns:
        xlsxファイルの内容をテキスト形式で表現した文字列
    """
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    
    lines: list[str] = []
    lines.append(f"=== xlsxファイル: {xlsx_path.name} ===")
    lines.append(f"シート名: {ws.title}")
    lines.append(f"データ範囲: A1:{ws.max_column}列 x {ws.max_row}行")
    lines.append("")
    
    # 全セルの内容を出力
    for row_idx in range(1, min(ws.max_row + 1, 100)):  # 最大100行まで
        row_values: list[str] = []
        for col_idx in range(1, min(ws.max_column + 1, 50)):  # 最大50列まで
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is not None:
                row_values.append(f"[{col_idx}]{cell_value}")
            else:
                row_values.append(f"[{col_idx}]")
        lines.append(f"行{row_idx}: " + " | ".join(row_values))
    
    if ws.max_row > 100:
        lines.append(f"... (残り{ws.max_row - 100}行省略)")
    
    return "\n".join(lines)


def _build_system_prompt() -> str:
    """システムプロンプトを構築する。"""
    return """あなたは開発プロジェクトの要員計画データを解析する専門家です。
Excelファイル（xlsx）の内容をテキスト形式で受け取り、以下の構造化されたJSON形式に変換してください。

出力するJSONスキーマ:
{
  "phases": [
    {
      "phase_name": "フェーズ名（例: 基本設計、詳細設計、開発・単体テスト）",
      "tasks": [
        {
          "task_name": "タスク名",
          "months": [
            {
              "year": 2024,
              "month": 4,
              "employee": 1.0,
              "pn": 2.0,
              "si": 3.0
            }
          ]
        }
      ]
    }
  ]
}

重要なルール:
1. フェーズ名は、Excelのデータ構造から推測してください（セクション見出し、グループ化された行など）
2. タスク名は、各行のタスク/作業項目を識別してください
3. 月別データは、年と月を正確に特定してください
   - 「年度」表記の場合: 4月-12月はその年、1月-3月は翌年として解釈
   - 例: 2024年度4月→year:2024,month:4、2024年度2月→year:2025,month:2
4. 工数の種類を識別してください:
   - employee（社員）: 「社員」「正社員」「Employee」などの表記
   - pn（PN）: 「PN」「パートナー」などの表記
   - si（SI）: 「SI」「SIer」「外注」「ベンダー」などの表記
5. 数値データがない場合は0.0としてください
6. フォーマットが不明確な場合は、最も合理的な解釈をしてください

必ずJSON形式のみで回答してください。説明文は不要です。"""


def _build_user_prompt(xlsx_text: str) -> str:
    """ユーザープロンプトを構築する。"""
    return f"""以下のExcelファイルの内容を解析し、要員計画データのJSON形式に変換してください。

{xlsx_text}

上記のデータを解析し、以下の形式のJSONを出力してください:
- phases配列にフェーズごとのデータ
- 各フェーズにtasks配列
- 各タスクにmonths配列（年月と3種類の工数）

JSONのみを出力してください。"""


def _extract_json_from_response(text: str) -> dict[str, Any]:
    """APIレスポンスからJSONを抽出する。"""
    # JSONブロックを抽出（```json ... ``` 形式）
    json_match = re.search(r"```json\s*([\s\S]*?)\s*```", text)
    if json_match:
        return json.loads(json_match.group(1))
    
    # ```なしのJSON
    json_match = re.search(r"\{[\s\S]*\}", text)
    if json_match:
        return json.loads(json_match.group(0))
    
    # 直接パース
    return json.loads(text)


def _convert_to_personnel_data(
    parsed_data: dict[str, Any],
    xlsx_path: Path,
) -> PersonnelData:
    """AIの出力をPersonnelDataモデルに変換する。"""
    phases_data = parsed_data.get("phases", [])
    phases: list[PersonnelPhaseData] = []
    
    for phase_dict in phases_data:
        phase_name = phase_dict.get("phase_name", "不明なフェーズ")
        tasks: list[PersonnelTaskData] = []
        
        for task_dict in phase_dict.get("tasks", []):
            task_name = task_dict.get("task_name", "不明なタスク")
            months: list[PersonnelMonthData] = []
            
            for month_dict in task_dict.get("months", []):
                year = int(month_dict.get("year", 2024))
                month = int(month_dict.get("month", 1))
                employee = float(month_dict.get("employee", 0.0))
                pn = float(month_dict.get("pn", 0.0))
                si = float(month_dict.get("si", 0.0))
                
                # 値が存在する場合のみ追加
                if any(v > 0 for v in [employee, pn, si]):
                    months.append(PersonnelMonthData(
                        year=year,
                        month=month,
                        employee=employee,
                        pn=pn,
                        si=si,
                    ))
            
            if months:
                tasks.append(PersonnelTaskData(
                    task_name=task_name,
                    months=months,
                ))
        
        if tasks:
            phases.append(PersonnelPhaseData(
                phase_name=phase_name,
                tasks=tasks,
            ))
    
    return PersonnelData(
        extracted_at=datetime.now(timezone.utc).isoformat(),
        source_path=str(xlsx_path),
        phases=phases,
    )


def _call_anthropic_api(
    xlsx_text: str,
    model: str | None,
    max_tokens: int | None,
) -> str:
    """Anthropic直接APIを呼び出す。"""
    try:
        from anthropic import Anthropic
    except ImportError as exc:
        msg = "anthropic パッケージが必要です。`pip install anthropic` を実行してください。"
        raise PersonnelAIParserConfigurationError(msg) from exc
    
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        raise PersonnelAIParserConfigurationError(
            "ANTHROPIC_API_KEY 環境変数が設定されていません"
        )
    
    model_name = model or os.getenv("ANTHROPIC_MODEL", DEFAULT_MODEL_ANTHROPIC)
    max_tokens_value = max_tokens or int(
        os.getenv("ANTHROPIC_MAX_TOKENS", str(DEFAULT_MAX_TOKENS))
    )
    temperature = float(os.getenv("ANTHROPIC_TEMPERATURE", "0.1"))
    
    client = Anthropic(api_key=api_key)
    
    _AI_LOGGER.info(
        "Calling Anthropic API: model=%s, max_tokens=%d",
        model_name,
        max_tokens_value,
    )
    
    response = client.messages.create(
        model=model_name,
        system=_build_system_prompt(),
        max_tokens=max_tokens_value,
        temperature=temperature,
        messages=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "text",
                        "text": _build_user_prompt(xlsx_text),
                    }
                ],
            }
        ],
    )
    
    # レスポンスからテキストを抽出
    text_parts = [
        block.text
        for block in response.content
        if getattr(block, "type", None) == "text"
    ]
    return "\n".join(text_parts)


def _call_bedrock_api(
    xlsx_text: str,
    model: str | None,
    max_tokens: int | None,
    region: str | None = None,
) -> str:
    """Amazon Bedrock経由でClaudeを呼び出す。"""
    try:
        import boto3
        from botocore.exceptions import ClientError, NoCredentialsError
    except ImportError as exc:
        msg = "boto3 パッケージが必要です。`pip install boto3` を実行してください。"
        raise PersonnelAIParserConfigurationError(msg) from exc
    
    # リージョン設定
    region_name = region or os.getenv("AWS_REGION", "us-east-1")
    
    # Bedrockクライアント作成
    try:
        client = boto3.client(
            "bedrock-runtime",
            region_name=region_name,
        )
    except NoCredentialsError as exc:
        raise PersonnelAIParserConfigurationError(
            "AWS認証情報が見つかりません。AWS_ACCESS_KEY_ID, AWS_SECRET_ACCESS_KEY を設定するか、"
            "AWS CLI でプロファイルを設定してください。"
        ) from exc
    
    model_id = model or os.getenv("BEDROCK_MODEL", DEFAULT_MODEL_BEDROCK)
    max_tokens_value = max_tokens or int(
        os.getenv("BEDROCK_MAX_TOKENS", str(DEFAULT_MAX_TOKENS))
    )
    temperature = float(os.getenv("BEDROCK_TEMPERATURE", "0.1"))
    
    _AI_LOGGER.info(
        "Calling Bedrock API: model=%s, region=%s, max_tokens=%d",
        model_id,
        region_name,
        max_tokens_value,
    )
    
    # Bedrock Converse APIを使用
    try:
        response = client.converse(
            modelId=model_id,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "text": _build_user_prompt(xlsx_text),
                        }
                    ],
                }
            ],
            system=[
                {
                    "text": _build_system_prompt(),
                }
            ],
            inferenceConfig={
                "maxTokens": max_tokens_value,
                "temperature": temperature,
            },
        )
    except NoCredentialsError as exc:
        raise PersonnelAIParserConfigurationError(
            "AWS認証情報が見つかりません。AWS_ACCESS_KEY_ID, AWS_SECRET_ACCESS_KEY を設定するか、"
            "AWS CLI でプロファイルを設定してください。"
        ) from exc
    except ClientError as exc:
        error_code = exc.response.get("Error", {}).get("Code", "Unknown")
        error_message = exc.response.get("Error", {}).get("Message", str(exc))
        if error_code == "UnrecognizedClientException":
            raise PersonnelAIParserConfigurationError(
                f"AWS認証情報が無効です。セキュリティトークンを確認してください: {error_message}"
            ) from exc
        if error_code == "AccessDeniedException":
            raise PersonnelAIParserConfigurationError(
                f"Bedrockへのアクセス権限がありません。IAMポリシーを確認してください: {error_message}"
            ) from exc
        if error_code == "ValidationException":
            raise PersonnelAIParserError(
                f"リクエストパラメータが無効です。モデルID '{model_id}' を確認してください: {error_message}"
            ) from exc
        raise PersonnelAIParserError(
            f"Bedrock API呼び出しエラー ({error_code}): {error_message}"
        ) from exc
    
    # レスポンスからテキストを抽出
    output = response.get("output", {})
    message = output.get("message", {})
    content = message.get("content", [])
    
    text_parts = [
        item.get("text", "")
        for item in content
        if "text" in item
    ]
    return "\n".join(text_parts)


def parse_personnel_xlsx_with_ai(
    xlsx_path: str | Path,
    *,
    model: str | None = None,
    max_tokens: int | None = None,
    backend: Literal["anthropic", "bedrock"] = "anthropic",
    region: str | None = None,
) -> PersonnelData:
    """AIを使用してxlsxファイルから工数データをパースする。
    
    Args:
        xlsx_path: 工数xlsxファイルのパス
        model: 使用するClaudeモデル
            - anthropic: デフォルト claude-sonnet-4-20250514
            - bedrock: デフォルト anthropic.claude-opus-4-5-20251101-v1:0
        max_tokens: 最大トークン数（デフォルト: 4096）
        backend: 使用するバックエンド（"anthropic" または "bedrock"）
        region: Bedrockのリージョン（デフォルト: us-east-1）
        
    Returns:
        PersonnelData: パースされた工数データ
        
    Raises:
        PersonnelAIParserConfigurationError: API設定エラー
        PersonnelAIParserError: パースエラー
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        msg = f"xlsxファイルが見つかりません: {xlsx_path}"
        raise FileNotFoundError(msg)
    
    # xlsxをテキストに変換
    xlsx_text = _xlsx_to_text(xlsx_path)
    _AI_LOGGER.info("xlsx converted to text: %d characters", len(xlsx_text))
    
    # バックエンドに応じてAPIを呼び出し
    if backend == "bedrock":
        response_text = _call_bedrock_api(xlsx_text, model, max_tokens, region)
    else:
        response_text = _call_anthropic_api(xlsx_text, model, max_tokens)
    
    _AI_LOGGER.info(
        "API response received: %d characters",
        len(response_text),
    )
    _AI_LOGGER.debug("Raw response: %s", response_text)
    
    # JSONをパース
    try:
        parsed_data = _extract_json_from_response(response_text)
    except json.JSONDecodeError as exc:
        msg = f"AIレスポンスのJSON解析に失敗しました: {exc}"
        _AI_LOGGER.error(msg)
        _AI_LOGGER.error("Response text: %s", response_text)
        raise PersonnelAIParserError(msg) from exc
    
    # PersonnelDataに変換
    personnel_data = _convert_to_personnel_data(parsed_data, xlsx_path)
    
    _AI_LOGGER.info(
        "Successfully parsed: %d phases, total tasks=%d",
        len(personnel_data.phases),
        sum(len(p.tasks) for p in personnel_data.phases),
    )
    
    return personnel_data


class MockPersonnelAIParser:
    """テスト用のモックAIパーサー。"""
    
    def __init__(self, response_data: dict[str, Any] | None = None) -> None:
        self._response_data = response_data or {"phases": []}
    
    def parse(self, xlsx_path: str | Path) -> PersonnelData:
        """モックパース。"""
        xlsx_path = Path(xlsx_path)
        return _convert_to_personnel_data(self._response_data, xlsx_path)


__all__ = [
    "PersonnelAIParserError",
    "PersonnelAIParserConfigurationError",
    "parse_personnel_xlsx_with_ai",
    "MockPersonnelAIParser",
]
